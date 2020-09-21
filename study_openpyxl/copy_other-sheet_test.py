from openpyxl import *

s_path = r'C:\Users\Administrator\Documents\GitHub\python-sql\python-excel\study_openpyxl\tt.xlsx'
d_path = r'C:\Users\Administrator\Documents\GitHub\python-sql\python-excel\study_openpyxl\copy_test.xlsx'

s_book = load_workbook(s_path)
s_sheet = s_book.get_sheet_by_name('base')

d_book = Workbook()
d_sheet = d_book.active

for i in range(1,s_sheet.max_row):
    for j in range(1,s_sheet.max_column):
        d_sheet.cell(row=i,column=j).value = list(s_sheet.rows)[i][j].value

d_sheet.title = 'test'
d_book.save(d_path)
#s_book.save(s_path)
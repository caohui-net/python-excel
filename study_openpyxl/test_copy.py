from openpyxl import *
testpath = r'C:\Users\Administrator\Documents\GitHub\python-sql\python-excel\study_openpyxl\2、铁八辅机BOM基础表.xlsx'

book = load_workbook(testpath)
sheet = book.active
print(book.sheetnames)
#sheet0 = book.create_sheet('',0)
#list_sheet = book.worksheets
#print(list_sheet[1].title)
#copy_sheet = book.copy_worksheet(book['base'])
#copy_sheet.title = 'base1'
#sheet = book['base1']
book.active = 2
sheet = book.active
print(sheet.title)
book.save(testpath)